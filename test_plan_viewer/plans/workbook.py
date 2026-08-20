"""Lossless XLSX import and export for Markdown test plans."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import zipfile
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_FORMULA
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException


PLAN_WORKBOOK_FORMAT_VERSION = 1
PLAN_WORKBOOK_SHEET = "测试计划"
PLAN_WORKBOOK_META_SHEET = "_meta"
PLAN_WORKBOOK_HEADERS = ("序号", "模块", "计划名称", "计划文件名")
PLAN_WORKBOOK_CONTENT_HEADER = "Markdown正文"
PLAN_WORKBOOK_CONTENT_CHUNK_CHARS = 30_000
PLAN_WORKBOOK_MAX_CONTENT_CHUNKS = 32
PLAN_WORKBOOK_MAX_ROWS = 5_000
PLAN_WORKBOOK_MAX_UPLOAD_BYTES = 20 * 1024 * 1024
PLAN_WORKBOOK_MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
PLAN_WORKBOOK_MAX_ARCHIVE_MEMBERS = 2_000
PLAN_WORKBOOK_CONFLICT_POLICIES = frozenset({"reject", "skip", "overwrite"})


class PlanWorkbookConflict(ValueError):
    """Raised when reject mode finds existing plans."""


@dataclass(frozen=True)
class PlanWorkbookDependencies:
    get_project_key: Callable[[], str]
    get_plan_file: Callable[[str, str], Path]
    validate_module_name: Callable[[str], str]
    validate_plan_filename: Callable[[str], str]
    sync_plan_asset: Callable[..., object]
    find_plan_asset: Callable[[Path], object]
    mark_plan_asset_deleted: Callable[[object], object]
    commit_removed_plan: Callable[[Path, str], object]
    current_timestamp: Callable[[], str]


def _set_text(cell, value):
    cell.value = str(value)
    cell.data_type = "s"


def _content_header(index):
    return (
        PLAN_WORKBOOK_CONTENT_HEADER
        if index == 1
        else f"{PLAN_WORKBOOK_CONTENT_HEADER}_{index}"
    )


def _split_content(content):
    max_chars = (
        PLAN_WORKBOOK_CONTENT_CHUNK_CHARS
        * PLAN_WORKBOOK_MAX_CONTENT_CHUNKS
    )
    if len(content) > max_chars:
        raise ValueError(
            f"测试计划正文超过 Excel 导出上限 {max_chars} 个字符。"
        )
    return [
        content[index : index + PLAN_WORKBOOK_CONTENT_CHUNK_CHARS]
        for index in range(0, len(content), PLAN_WORKBOOK_CONTENT_CHUNK_CHARS)
    ] or [""]


def _validate_archive(data):
    if not data:
        raise ValueError("导入 Excel 不能为空。")
    if len(data) > PLAN_WORKBOOK_MAX_UPLOAD_BYTES:
        raise ValueError(
            "导入 Excel 不能超过 "
            f"{PLAN_WORKBOOK_MAX_UPLOAD_BYTES // 1024 // 1024}MB。"
        )
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > PLAN_WORKBOOK_MAX_ARCHIVE_MEMBERS:
                raise ValueError("导入 Excel 包含过多内部文件。")
            if sum(member.file_size for member in members) > PLAN_WORKBOOK_MAX_UNCOMPRESSED_BYTES:
                raise ValueError("导入 Excel 解压后不能超过 100MB。")
            for member in members:
                normalized_name = member.filename.replace("\\", "/")
                if normalized_name.startswith("/") or ".." in Path(
                    normalized_name
                ).parts:
                    raise ValueError("导入 Excel 包含非法内部路径。")
                name = normalized_name.lower()
                if member.flag_bits & 0x1:
                    raise ValueError("不支持加密的 Excel 文件。")
                if "vbaproject.bin" in name:
                    raise ValueError("不支持包含宏的 Excel 文件。")
                if name.startswith("xl/externallinks/"):
                    raise ValueError("不支持包含外部链接的 Excel 文件。")
                if name.endswith(".rels"):
                    relationship_xml = archive.read(member).lower()
                    if (
                        b'targetmode="external"' in relationship_xml
                        or b"targetmode='external'" in relationship_xml
                    ):
                        raise ValueError("不支持包含外部链接的 Excel 文件。")
    except zipfile.BadZipFile as exc:
        raise ValueError("导入文件不是合法的 .xlsx。") from exc


class PlanWorkbookService:
    def __init__(self, dependencies):
        if not isinstance(dependencies, PlanWorkbookDependencies):
            raise TypeError("dependencies must be PlanWorkbookDependencies")
        self.dependencies = dependencies

    def export(self, selections):
        rows = self._load_export_rows(selections)
        exported_at = self.dependencies.current_timestamp()
        max_chunks = max(len(row["chunks"]) for row in rows)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = PLAN_WORKBOOK_SHEET
        sheet.sheet_view.showGridLines = False
        headers = [
            *PLAN_WORKBOOK_HEADERS,
            *[_content_header(index) for index in range(1, max_chunks + 1)],
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, row in enumerate(rows, start=1):
            sheet.append([row_index, "", "", "", *([""] * max_chunks)])
            target_row = sheet.max_row
            _set_text(sheet.cell(target_row, 2), row["module_name"])
            _set_text(sheet.cell(target_row, 3), row["plan_name"])
            _set_text(sheet.cell(target_row, 4), row["plan_filename"])
            for chunk_index, chunk in enumerate(row["chunks"], start=5):
                _set_text(sheet.cell(target_row, chunk_index), chunk)
            for cell in sheet[target_row]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 24
        sheet.column_dimensions["A"].width = 9
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 32
        sheet.column_dimensions["D"].width = 34
        for column_index in range(5, 5 + max_chunks):
            sheet.column_dimensions[
                sheet.cell(1, column_index).column_letter
            ].width = 80

        meta = workbook.create_sheet(PLAN_WORKBOOK_META_SHEET)
        meta.sheet_state = "hidden"
        metadata = (
            ("format_version", PLAN_WORKBOOK_FORMAT_VERSION),
            ("project_key", self.dependencies.get_project_key()),
            ("exported_at", exported_at),
        )
        for key, value in metadata:
            meta.append([key, value])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        project_key = self.dependencies.get_project_key() or "project"
        filename = (
            f"测试计划-{project_key}-"
            f"{exported_at}.xlsx"
        )
        return buffer, filename

    def _load_export_rows(self, selections):
        if not isinstance(selections, list) or not selections:
            raise ValueError("请至少选择一个测试计划。")
        if len(selections) > PLAN_WORKBOOK_MAX_ROWS:
            raise ValueError(f"一次最多导出 {PLAN_WORKBOOK_MAX_ROWS} 条测试计划。")
        rows = []
        seen = set()
        for selection in selections:
            if not isinstance(selection, dict):
                raise ValueError("导出计划选择格式错误。")
            module_name = self.dependencies.validate_module_name(
                str(selection.get("module_name") or "").strip()
            )
            plan_filename = self.dependencies.validate_plan_filename(
                str(selection.get("plan_filename") or "").strip()
            )
            key = (module_name, plan_filename)
            if key in seen:
                raise ValueError(f"导出列表包含重复测试计划：{module_name}/{plan_filename}")
            seen.add(key)
            path = self.dependencies.get_plan_file(module_name, plan_filename)
            if not path.is_file():
                raise FileNotFoundError(f"测试计划不存在：{module_name}/{plan_filename}")
            content = path.read_text(encoding="utf-8")
            rows.append(
                {
                    "module_name": module_name,
                    "plan_name": path.stem,
                    "plan_filename": plan_filename,
                    "chunks": _split_content(content),
                }
            )
        return rows

    def import_bytes(self, data, conflict_policy="reject"):
        conflict_policy = str(conflict_policy or "reject").strip().lower()
        if conflict_policy not in PLAN_WORKBOOK_CONFLICT_POLICIES:
            raise ValueError("同名处理策略必须是 reject、skip 或 overwrite。")
        _validate_archive(data)
        rows = self._parse_rows(data)
        prepared = []
        conflicts = []
        for row in rows:
            target = self.dependencies.get_plan_file(
                row["module_name"], row["plan_filename"]
            )
            exists = target.exists()
            if exists and conflict_policy == "reject":
                conflicts.append(f"{row['module_name']}/{row['plan_filename']}")
            action = "skipped" if exists and conflict_policy == "skip" else (
                "overwritten" if exists else "created"
            )
            prepared.append({**row, "target": target, "exists": exists, "action": action})
        if conflicts:
            preview = "、".join(conflicts[:20])
            suffix = "……" if len(conflicts) > 20 else ""
            raise PlanWorkbookConflict(f"以下测试计划已存在：{preview}{suffix}")

        applied = []
        results = []
        try:
            for row in prepared:
                if row["action"] == "skipped":
                    results.append(self._result_item(row))
                    continue
                target = row["target"]
                backup = target.read_bytes() if row["exists"] else None
                row["parent_existed"] = target.parent.exists()
                applied.append((row, backup))
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_text(row["markdown"], encoding="utf-8", newline="")
                self.dependencies.sync_plan_asset(
                    row["module_name"],
                    target,
                    change_source="import",
                    message=(
                        "excel import: "
                        f"{row['module_name']}/{row['plan_filename']}"
                    ),
                )
                results.append(self._result_item(row))
        except Exception as exc:
            rollback_errors = self._rollback(applied)
            detail = f"；回滚异常：{'；'.join(rollback_errors)}" if rollback_errors else ""
            raise RuntimeError(f"导入测试计划失败，已恢复当前文件：{exc}{detail}") from exc

        counts = {
            action: sum(item["action"] == action for item in results)
            for action in ("created", "overwritten", "skipped")
        }
        return {
            **counts,
            "total": len(results),
            "items": results,
            "error": None,
        }

    @staticmethod
    def _result_item(row):
        return {
            "module_name": row["module_name"],
            "plan_filename": row["plan_filename"],
            "action": row["action"],
        }

    def _rollback(self, applied):
        errors = []
        for row, backup in reversed(applied):
            target = row["target"]
            try:
                if backup is None:
                    target.unlink(missing_ok=True)
                    asset = self.dependencies.find_plan_asset(target)
                    if asset:
                        self.dependencies.mark_plan_asset_deleted(asset)
                    self.dependencies.commit_removed_plan(
                        target,
                        "excel import rollback: "
                        f"{row['module_name']}/{row['plan_filename']}",
                    )
                    if not row.get("parent_existed"):
                        target.parent.rmdir()
                else:
                    target.write_bytes(backup)
                    self.dependencies.sync_plan_asset(
                        row["module_name"],
                        target,
                        change_source="import",
                        message=(
                            "excel import rollback: "
                            f"{row['module_name']}/{row['plan_filename']}"
                        ),
                    )
            except Exception as rollback_exc:
                errors.append(
                    f"{row['module_name']}/{row['plan_filename']}：{rollback_exc}"
                )
        return errors

    def _parse_rows(self, data):
        try:
            workbook = load_workbook(
                BytesIO(data),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValueError("无法读取导入 Excel。") from exc
        try:
            self._validate_metadata(workbook)
            if set(workbook.sheetnames) != {
                PLAN_WORKBOOK_SHEET,
                PLAN_WORKBOOK_META_SHEET,
            }:
                raise ValueError("导入 Excel 包含未知工作表。")
            if PLAN_WORKBOOK_SHEET not in workbook.sheetnames:
                raise ValueError(f"导入 Excel 缺少“{PLAN_WORKBOOK_SHEET}”工作表。")
            sheet = workbook[PLAN_WORKBOOK_SHEET]
            header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
            if not header_cells:
                raise ValueError("导入 Excel 缺少表头。")
            if any(cell.data_type == TYPE_FORMULA for cell in header_cells):
                raise ValueError("导入 Excel 表头不能包含公式。")
            headers = [str(cell.value or "").strip() for cell in header_cells]
            if len(headers) > (
                len(PLAN_WORKBOOK_HEADERS) + PLAN_WORKBOOK_MAX_CONTENT_CHUNKS
            ):
                raise ValueError("导入 Excel 包含未知列。")
            expected_prefix = [*PLAN_WORKBOOK_HEADERS, PLAN_WORKBOOK_CONTENT_HEADER]
            if headers[: len(expected_prefix)] != expected_prefix:
                raise ValueError("导入 Excel 表头不符合平台格式。")
            extra_content_headers = headers[len(expected_prefix) :]
            content_columns = 1 + len(extra_content_headers)
            for index, header in enumerate(extra_content_headers, start=2):
                if header != _content_header(index):
                    raise ValueError("Markdown 正文分片列必须连续编号。")
            if content_columns > PLAN_WORKBOOK_MAX_CONTENT_CHUNKS:
                raise ValueError("Markdown 正文分片列超过平台上限。")

            parsed = []
            seen = set()
            max_column = 4 + content_columns
            for row_number, cells in enumerate(
                sheet.iter_rows(min_row=2, max_col=max_column), start=2
            ):
                if all(cell.value in (None, "") for cell in cells):
                    continue
                if len(parsed) >= PLAN_WORKBOOK_MAX_ROWS:
                    raise ValueError(f"一次最多导入 {PLAN_WORKBOOK_MAX_ROWS} 条测试计划。")
                for cell in cells[:max_column]:
                    if cell.data_type == TYPE_FORMULA:
                        raise ValueError(f"第 {row_number} 行不能包含公式。")
                values = [cell.value for cell in cells]
                module_name = self._required_text(values[1], row_number, "模块")
                plan_name = self._optional_text(values[2], row_number, "计划名称")
                plan_filename = self._required_text(values[3], row_number, "计划文件名")
                try:
                    module_name = self.dependencies.validate_module_name(module_name)
                except ValueError as exc:
                    raise ValueError(
                        f"第 {row_number} 行模块无效：{exc}"
                    ) from exc
                try:
                    plan_filename = self.dependencies.validate_plan_filename(
                        plan_filename
                    )
                except ValueError as exc:
                    raise ValueError(
                        f"第 {row_number} 行计划文件名无效：{exc}"
                    ) from exc
                if plan_name and plan_name != Path(plan_filename).stem:
                    raise ValueError(
                        f"第 {row_number} 行计划名称必须等于计划文件名去掉 .md 后的名称。"
                    )
                chunks = []
                for column_index, value in enumerate(values[4:max_column], start=1):
                    chunk = self._optional_text(
                        value, row_number, _content_header(column_index)
                    )
                    if len(chunk) > PLAN_WORKBOOK_CONTENT_CHUNK_CHARS:
                        raise ValueError(
                            f"第 {row_number} 行 {_content_header(column_index)} 超过字符上限。"
                        )
                    chunks.append(chunk)
                markdown = "".join(chunks)
                if not markdown:
                    raise ValueError(f"第 {row_number} 行 Markdown正文不能为空。")
                key = (module_name, plan_filename)
                if key in seen:
                    raise ValueError(
                        f"Excel 包含重复测试计划：{module_name}/{plan_filename}"
                    )
                seen.add(key)
                parsed.append(
                    {
                        "module_name": module_name,
                        "plan_filename": plan_filename,
                        "markdown": markdown,
                    }
                )
            if not parsed:
                raise ValueError("导入 Excel 中没有测试计划。")
            return parsed
        finally:
            workbook.close()

    @staticmethod
    def _required_text(value, row_number, label):
        text = PlanWorkbookService._optional_text(value, row_number, label)
        if not text:
            raise ValueError(f"第 {row_number} 行{label}不能为空。")
        return text

    @staticmethod
    def _optional_text(value, row_number, label):
        if value is None:
            return ""
        if not isinstance(value, str):
            raise ValueError(f"第 {row_number} 行{label}必须是文本。")
        return value

    @staticmethod
    def _validate_metadata(workbook):
        if PLAN_WORKBOOK_META_SHEET not in workbook.sheetnames:
            raise ValueError("导入 Excel 缺少格式版本信息。")
        meta = workbook[PLAN_WORKBOOK_META_SHEET]
        values = {
            str(row[0].value or "").strip(): row[1].value
            for row in meta.iter_rows(min_row=1, max_col=2)
            if row[0].value
        }
        if any(
            cell.data_type == TYPE_FORMULA
            for row in meta.iter_rows()
            for cell in row
        ):
            raise ValueError("导入 Excel 元数据不能包含公式。")
        try:
            version = int(values.get("format_version"))
        except (TypeError, ValueError) as exc:
            raise ValueError("导入 Excel 格式版本无效。") from exc
        if version != PLAN_WORKBOOK_FORMAT_VERSION:
            raise ValueError("不支持的测试计划 Excel 格式版本。")

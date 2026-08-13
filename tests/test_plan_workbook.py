import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from test_plan_viewer.artifacts.naming import (
    validate_module_name,
    validate_plan_filename,
)
from test_plan_viewer.plans import workbook as workbook_module
from test_plan_viewer.plans.workbook import (
    PLAN_WORKBOOK_META_SHEET,
    PLAN_WORKBOOK_SHEET,
    PlanWorkbookConflict,
    PlanWorkbookDependencies,
    PlanWorkbookService,
)


class PlanWorkbookHarness:
    def __init__(self, root, *, sync_plan_asset=None):
        self.root = Path(root)
        self.sync_calls = []
        self.deleted_assets = []
        self.assets = {}
        self._sync_plan_asset = sync_plan_asset
        self.service = PlanWorkbookService(
            PlanWorkbookDependencies(
                get_project_key=lambda: "project-alpha",
                get_plan_file=self.get_plan_file,
                validate_module_name=validate_module_name,
                validate_plan_filename=validate_plan_filename,
                sync_plan_asset=self.sync_plan_asset,
                find_plan_asset=lambda path: self.assets.get(str(path)),
                mark_plan_asset_deleted=self.deleted_assets.append,
                commit_removed_plan=lambda _path, _message: None,
                current_timestamp=lambda: "20260806-120000",
            )
        )

    def get_plan_file(self, module_name, plan_filename):
        module_name = validate_module_name(module_name)
        plan_filename = validate_plan_filename(plan_filename)
        return self.root / module_name / plan_filename

    def sync_plan_asset(self, module_name, path, **kwargs):
        call = (module_name, Path(path), kwargs)
        self.sync_calls.append(call)
        if self._sync_plan_asset:
            return self._sync_plan_asset(self, *call)
        asset = {"path": str(path)}
        self.assets[str(path)] = asset
        return asset

    def write(self, module_name, filename, content):
        path = self.get_plan_file(module_name, filename)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")
        return path

    def export(self, *items):
        buffer, _filename = self.service.export(
            [
                {"module_name": module, "plan_filename": filename}
                for module, filename in items
            ]
        )
        return buffer.getvalue()


def mutate_workbook(data, callback):
    workbook = load_workbook(BytesIO(data))
    callback(workbook)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def append_archive_member(data, name, content=b"blocked"):
    output = BytesIO(data)
    with ZipFile(output, "a", compression=ZIP_DEFLATED) as archive:
        archive.writestr(name, content)
    return output.getvalue()


class PlanWorkbookRoundTripTests(unittest.TestCase):
    def test_unicode_multiline_long_markdown_round_trips_losslessly(self):
        content = "=不是公式\n# 中文计划\n\n" + ("多行内容🙂\n" * 8_000)
        with tempfile.TemporaryDirectory() as source_dir, tempfile.TemporaryDirectory() as target_dir:
            source = PlanWorkbookHarness(source_dir)
            source.write("贷款管理", "核心回归.md", content)
            source.write("用户管理", "权限检查.md", "# 权限\n\n- 管理员\n")

            data = source.export(
                ("贷款管理", "核心回归.md"),
                ("用户管理", "权限检查.md"),
            )
            workbook = load_workbook(BytesIO(data), data_only=False)
            self.assertEqual(workbook.sheetnames, [PLAN_WORKBOOK_SHEET, PLAN_WORKBOOK_META_SHEET])
            self.assertEqual(workbook[PLAN_WORKBOOK_META_SHEET].sheet_state, "hidden")
            sheet = workbook[PLAN_WORKBOOK_SHEET]
            self.assertEqual(sheet["E1"].value, "Markdown正文")
            self.assertEqual(sheet["F1"].value, "Markdown正文_2")
            self.assertEqual(sheet["E2"].data_type, "s")
            self.assertEqual(sheet["E2"].value[0], "=")
            workbook.close()

            target = PlanWorkbookHarness(target_dir)
            result = target.service.import_bytes(data)

            self.assertEqual(result["created"], 2)
            self.assertEqual(result["overwritten"], 0)
            self.assertEqual(result["skipped"], 0)
            self.assertEqual(result["total"], 2)
            self.assertEqual(
                target.get_plan_file("贷款管理", "核心回归.md").read_text(encoding="utf-8"),
                content,
            )
            self.assertTrue(
                all(call[2]["change_source"] == "import" for call in target.sync_calls)
            )

    def test_export_rejects_empty_duplicate_missing_and_oversized_selections(self):
        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root)
            harness.write("模块", "计划.md", "正文")
            with self.assertRaisesRegex(ValueError, "至少选择"):
                harness.service.export([])
            with self.assertRaisesRegex(ValueError, "重复"):
                harness.service.export(
                    [
                        {"module_name": "模块", "plan_filename": "计划.md"},
                        {"module_name": "模块", "plan_filename": "计划.md"},
                    ]
                )
            with self.assertRaisesRegex(FileNotFoundError, "不存在"):
                harness.service.export(
                    [{"module_name": "模块", "plan_filename": "缺失.md"}]
                )
            harness.write(
                "模块",
                "超长.md",
                "字" * (
                    workbook_module.PLAN_WORKBOOK_CONTENT_CHUNK_CHARS
                    * workbook_module.PLAN_WORKBOOK_MAX_CONTENT_CHUNKS
                    + 1
                ),
            )
            with self.assertRaisesRegex(ValueError, "正文超过"):
                harness.export(("模块", "超长.md"))


class PlanWorkbookImportTests(unittest.TestCase):
    def setUp(self):
        self.source_directory = tempfile.TemporaryDirectory()
        self.source = PlanWorkbookHarness(self.source_directory.name)
        self.source.write("模块甲", "计划甲.md", "新正文甲")
        self.source.write("模块乙", "计划乙.md", "新正文乙")
        self.data = self.source.export(
            ("模块甲", "计划甲.md"),
            ("模块乙", "计划乙.md"),
        )

    def tearDown(self):
        self.source_directory.cleanup()

    def test_conflict_policies_reject_skip_and_overwrite(self):
        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root)
            existing = harness.write("模块甲", "计划甲.md", "旧正文")

            with self.assertRaises(PlanWorkbookConflict):
                harness.service.import_bytes(self.data, "reject")
            self.assertEqual(existing.read_text(encoding="utf-8"), "旧正文")
            self.assertFalse(harness.get_plan_file("模块乙", "计划乙.md").exists())

            skipped = harness.service.import_bytes(self.data, "skip")
            self.assertEqual(skipped["created"], 1)
            self.assertEqual(skipped["skipped"], 1)
            self.assertEqual(existing.read_text(encoding="utf-8"), "旧正文")

            overwritten = harness.service.import_bytes(self.data, "overwrite")
            self.assertEqual(overwritten["overwritten"], 2)
            self.assertEqual(existing.read_text(encoding="utf-8"), "新正文甲")
            self.assertEqual(harness.sync_calls[-1][2]["change_source"], "import")

    def test_format_path_duplicate_and_formula_validation(self):
        cases = {
            "表头": lambda wb: setattr(wb[PLAN_WORKBOOK_SHEET]["A1"], "value", "错误"),
            "计划名称": lambda wb: setattr(wb[PLAN_WORKBOOK_SHEET]["C2"], "value", "别名"),
            "公式": lambda wb: setattr(wb[PLAN_WORKBOOK_SHEET]["A2"], "value", "=1+1"),
            "模块无效": lambda wb: setattr(
                wb[PLAN_WORKBOOK_SHEET]["B2"], "value", "../模块"
            ),
            "重复": lambda wb: wb[PLAN_WORKBOOK_SHEET].append(
                [2, "模块甲", "计划甲", "计划甲.md", "重复正文"]
            ),
            "格式版本": lambda wb: setattr(wb[PLAN_WORKBOOK_META_SHEET]["B1"], "value", 99),
            "未知工作表": lambda wb: wb.create_sheet("额外数据"),
        }
        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root)
            for expected, mutation in cases.items():
                with self.subTest(expected=expected):
                    data = mutate_workbook(self.data, mutation)
                    with self.assertRaisesRegex(ValueError, expected):
                        harness.service.import_bytes(data)

    def test_compressed_and_uncompressed_size_limits_are_enforced(self):
        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root)
            with patch.object(workbook_module, "PLAN_WORKBOOK_MAX_UPLOAD_BYTES", 10):
                with self.assertRaisesRegex(ValueError, "不能超过"):
                    harness.service.import_bytes(b"x" * 11)
            with patch.object(
                workbook_module,
                "PLAN_WORKBOOK_MAX_UNCOMPRESSED_BYTES",
                100,
            ):
                with self.assertRaisesRegex(ValueError, "解压后"):
                    harness.service.import_bytes(self.data)

    def test_row_limit_and_dangerous_archive_members_are_rejected(self):
        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root)
            with patch.object(workbook_module, "PLAN_WORKBOOK_MAX_ROWS", 1):
                with self.assertRaisesRegex(ValueError, "最多导入"):
                    harness.service.import_bytes(self.data)

            dangerous_members = {
                "../escape.xml": "非法内部路径",
                "xl/vbaProject.bin": "宏",
                "xl/externalLinks/externalLink1.xml": "外部链接",
                "xl/worksheets/_rels/sheet1.xml.rels": "外部链接",
            }
            for member, expected in dangerous_members.items():
                with self.subTest(member=member):
                    content = (
                        b'<Relationship TargetMode="External"/>'
                        if member.endswith(".rels")
                        else b"blocked"
                    )
                    data = append_archive_member(self.data, member, content)
                    with self.assertRaisesRegex(ValueError, expected):
                        harness.service.import_bytes(data)

    def test_sync_failure_restores_all_current_files_and_new_module_directory(self):
        def fail_second_sync(harness, _module_name, path, _kwargs):
            if len(harness.sync_calls) == 2:
                harness.assets[str(path)] = {"path": str(path)}
                raise RuntimeError("asset sync failed")
            return {"path": str(path)}

        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root, sync_plan_asset=fail_second_sync)
            existing = harness.write("模块甲", "计划甲.md", "旧正文")
            new_module = harness.root / "模块乙"

            with self.assertRaisesRegex(RuntimeError, "已恢复当前文件"):
                harness.service.import_bytes(self.data, "overwrite")

            self.assertEqual(existing.read_text(encoding="utf-8"), "旧正文")
            self.assertFalse(new_module.exists())
            self.assertEqual(len(harness.deleted_assets), 1)

    def test_file_write_failure_restores_existing_content(self):
        with tempfile.TemporaryDirectory() as root:
            harness = PlanWorkbookHarness(root)
            existing = harness.write("模块甲", "计划甲.md", "旧正文")
            original_write_text = Path.write_text

            def fail_import_write(path, data, *args, **kwargs):
                if path == existing and data == "新正文甲":
                    path.write_bytes(b"partial")
                    raise OSError("disk full")
                return original_write_text(path, data, *args, **kwargs)

            with patch.object(Path, "write_text", fail_import_write):
                with self.assertRaisesRegex(RuntimeError, "已恢复当前文件"):
                    harness.service.import_bytes(self.data, "overwrite")

            self.assertEqual(existing.read_text(encoding="utf-8"), "旧正文")


if __name__ == "__main__":
    unittest.main()
